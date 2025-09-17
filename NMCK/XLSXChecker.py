# scan_nmck_orders_in_excel.py
import os
import re
import csv
from typing import Dict, List, Tuple, Set

# --- НАСТРОЙКИ ---
TARGET_DIR = r"C:\Users\Sergey\MAGA\Kimch\НМЦК_сборка\223- фз"
RECURSIVE = False  # при необходимости можно поставить True
WRITE_SNIPPETS = True  # писать в CSV фрагменты текста вокруг совпадений
TARGET_ARTICLES = {"22", "34"}  # какие статьи считаем целевыми
# --- ПАТТЕРНЫ ПОИСКА ---
# ст.22 / статья 22
RE_ART_ANY = re.compile(
    r"(?<!\w)(?:[сc][тt](?:\s*\.)? | стать(?:я|и|е|ю|ёй|ей))\s*([0-9]{1,3}(?:\.\d+)*)\b",
    re.IGNORECASE | re.UNICODE | re.VERBOSE
)

# приказы 567/639: не часть большего числа слева и
# НЕ должны быть дальше цифры (с учётом пробелов/.,,)
RE_ORDERS = re.compile(r"(?<!\d)(567|639)(?![\s.,]*\d)", re.UNICODE)

# номер закупки из имени: приоритетно "Закупка № <19 цифр>", запасной — любой 19-значный
RE_NUM_FROM_TITLE = re.compile(r"Закупка\s*№\s*(\d{11})", re.UNICODE | re.IGNORECASE)
RE_ANY_19_DIGITS = re.compile(r"(?<!\d)(\d{11})(?!\d)", re.UNICODE)

def extract_purchase_number_from_filename(filename: str) -> str:
    m = RE_NUM_FROM_TITLE.search(filename)
    if m:
        return m.group(1)
    m2 = RE_ANY_19_DIGITS.search(filename)
    return m2.group(1) if m2 else "UNKNOWN"

def read_excel_cells(filepath: str) -> List[str]:
    """
    Возвращает список текстов всех ячеек книги (включая все листы) как строки.
    Сначала пробует pandas, затем openpyxl (для .xlsx).
    Для .xls без xlrd выдаст предупреждение.
    """
    texts: List[str] = []
    ext = os.path.splitext(filepath.lower())[1]

    # 1) Пробуем pandas (универсально)
    try:
        import pandas as pd
        dfs = pd.read_excel(filepath, sheet_name=None, dtype=str)
        for _, df in dfs.items():
            # преобразуем все значения в строки, заполнители NaN -> ''
            vals = df.astype(str).values.ravel().tolist()
            texts.extend([v for v in vals if v and v != "nan"])
        return texts
    except Exception as e_pd:
        # если .xlsx — попробуем openpyxl
        if ext == ".xlsx":
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath, data_only=True, read_only=True)
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        for v in row:
                            if v is None:
                                continue
                            texts.append(str(v))
                return texts
            except Exception as e_xlsx:
                print(f"[WARN] Не удалось прочитать xlsx через pandas/openpyxl: {filepath}\n  {e_xlsx}")
                return texts
        # если .xls — возможна проблема с xlrd
        if ext == ".xls":
            try:
                import xlrd  # требуется xlrd<=1.2.0
                book = xlrd.open_workbook(filepath)
                for si in range(book.nsheets):
                    sh = book.sheet_by_index(si)
                    for r in range(sh.nrows):
                        for c in range(sh.ncols):
                            v = sh.cell_value(r, c)
                            if v != "":
                                texts.append(str(v))
                return texts
            except Exception as e_xls:
                print(f"[WARN] Не удалось прочитать xls (нужен xlrd, лучше версии 1.2.0): {filepath}\n  {e_xls}")
                return texts

        # прочие ошибки
        print(f"[WARN] Ошибка чтения Excel: {filepath}\n  {e_pd}")
        return texts

def find_matches(texts: List[str]) -> Tuple[Set[str], Dict[str, List[str]], Dict[str, List[str]]]:
    """
    Возвращает:
      - found_articles: множество корней статей {'22', '34'} найденных в тексте
      - orders_snips: {'567': [сниппеты], '639': [сниппеты]}
      - art_snips: {'22': [сниппеты], '34': [сниппеты]}
    """
    found_articles: Set[str] = set()
    orders_snips: Dict[str, List[str]] = {"567": [], "639": []}
    art_snips: Dict[str, List[str]] = {}

    def snippet(s: str, m: re.Match, radius: int = 40) -> str:
        if not WRITE_SNIPPETS:
            return ""
        start = max(0, m.start() - radius)
        end = min(len(s), m.end() + radius)
        return s[start:end].replace("\n", " ")

    for raw in texts:
        s = _normalize_ws(str(raw))

        # Статьи (22/34 + подпункты типа 22.1)
        for m in RE_ART_ANY.finditer(s):
            art_num = m.group(1)
            root = art_num.split('.')[0]  # '22' из '22.1'
            if root in TARGET_ARTICLES:
                found_articles.add(root)
                art_snips.setdefault(root, []).append(snippet(s, m))

        # Приказы 567/639
        for m in RE_ORDERS.finditer(s):
            ord_num = m.group(1)
            orders_snips.setdefault(ord_num, []).append(snippet(s, m))

    # Уберём пустые ключи
    orders_snips = {k: v for k, v in orders_snips.items() if v}
    art_snips = {k: v for k, v in art_snips.items() if v}
    return found_articles, orders_snips, art_snips


def list_excel_files(root: str, recursive: bool = False) -> List[str]:
    res = []
    if recursive:
        for dp, _, fns in os.walk(root):
            for fn in fns:
                if fn.lower().endswith((".xlsx", ".xls")):
                    res.append(os.path.join(dp, fn))
    else:
        if not os.path.isdir(root):
            return []
        for fn in os.listdir(root):
            if fn.lower().endswith((".xlsx", ".xls")):
                res.append(os.path.join(root, fn))
    return res

def write_csv(path: str, rows: List[List[str]], header: List[str]):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(header)
        w.writerows(rows)
def _normalize_ws(s: str) -> str:
    # NBSP, thin space и т.п. -> обычный пробел
    return s.replace("\u00A0", " ").replace("\u2009", " ").replace("\u202F", " ")

def main():
    # Если нужно строго "только 639 и без 567" — поставь True
    STRICT_639_ONLY = False

    files = list_excel_files(TARGET_DIR, RECURSIVE)
    if not files:
        print(f"В папке нет Excel-файлов: {TARGET_DIR}")
        return

    rows_found_any: List[List[str]] = []   # статья или приказ, или оба
    rows_none: List[List[str]] = []        # ни статьи, ни приказа
    rows_only_639: List[List[str]] = []    # отдельный отчёт: все, где есть 639

    total_files = len(files)
    xlsx_total = sum(1 for p in files if p.lower().endswith(".xlsx"))
    xls_total  = sum(1 for p in files if p.lower().endswith(".xls"))
    xlsx_read_ok = 0
    xls_read_ok  = 0

    # счётчики по категориям
    cnt_both = 0
    cnt_orders_only = 0
    cnt_articles_only = 0
    cnt_none = 0
    cnt_with_639 = 0

    for path in files:
        fname = os.path.basename(path)
        purchase = extract_purchase_number_from_filename(fname)

        texts = read_excel_cells(path)

        # учёт успешности чтения по расширению
        low = path.lower()
        if low.endswith(".xlsx") and texts:
            xlsx_read_ok += 1
        if low.endswith(".xls") and texts:
            xls_read_ok += 1

        # поиск
        found_articles, orders_snips, art_snips = find_matches(texts)

        # агрегаты для CSV
        articles_found = "|".join(sorted(found_articles)) if found_articles else ""
        orders_found   = "|".join(sorted(orders_snips.keys())) if orders_snips else ""

        # сниппеты (по 1–2 на тип)
        if WRITE_SNIPPETS:
            article_snippets = " || ".join(
                f"{k}: " + " | ".join(v[:2]) for k, v in sorted(art_snips.items())
            ) if art_snips else ""
            order_snippets = " || ".join(
                f"{k}: " + " | ".join(v[:2]) for k, v in sorted(orders_snips.items())
            ) if orders_snips else ""
        else:
            article_snippets = ""
            order_snippets = ""

        # статус и распределение по файлам (found_any / not_found_any)
        if found_articles and orders_snips:
            status = "orders+article"
            cnt_both += 1
            rows_found_any.append([
                purchase, articles_found, orders_found, status,
                fname, path, article_snippets, order_snippets
            ])
        elif orders_snips:
            status = "orders"
            cnt_orders_only += 1
            rows_found_any.append([
                purchase, articles_found, orders_found, status,
                fname, path, article_snippets, order_snippets
            ])
        elif found_articles:
            status = "article"
            cnt_articles_only += 1
            rows_found_any.append([
                purchase, articles_found, orders_found, status,
                fname, path, article_snippets, order_snippets
            ])
        else:
            status = "missing"
            cnt_none += 1
            rows_none.append([purchase, fname, path])

        # --- ОТЧЁТ ТОЛЬКО ПО 639 ---
        has_639 = "639" in orders_snips
        if STRICT_639_ONLY:
            has_639 = has_639 and ("567" not in orders_snips)

        if has_639:
            cnt_with_639 += 1
            order_639_snippets = " | ".join(orders_snips.get("639", [])[:2]) if WRITE_SNIPPETS else ""
            rows_only_639.append([
                purchase, articles_found, orders_found, status,
                fname, path, order_639_snippets
            ])

    # пишем отчёты
    base = os.path.dirname(os.path.abspath(__file__))
    write_csv(os.path.join(base, "found_any.csv"),
              rows_found_any,
              header=["purchase_number", "articles_found", "orders_found", "status",
                      "file_name", "file_path", "article_snippets", "order_snippets"])
    write_csv(os.path.join(base, "not_found_any.csv"),
              rows_none,
              header=["purchase_number", "file_name", "file_path"])
    write_csv(os.path.join(base, "only_639_excel.csv"),
              rows_only_639,
              header=["purchase_number", "articles_found", "orders_found", "status",
                      "file_name", "file_path", "order_639_snippets"])

    # итоги в консоль
    print(f"Всего файлов (xls/xlsx): {total_files}")
    print(f"XLSX: прочитано {xlsx_read_ok} из {xlsx_total}")
    print(f"XLS : прочитано {xls_read_ok} из {xls_total}")
    print(f"Найдено ИЛИ статья, ИЛИ приказ, ИЛИ оба (found_any): {len(rows_found_any)}")
    print(f"  — только приказ: {cnt_orders_only}")
    print(f"  — только статья: {cnt_articles_only}")
    print(f"  — статья+приказ: {cnt_both}")
    print(f"Ничего не найдено (not_found_any): {cnt_none}")
    print(f"Документов с приказом 639: {cnt_with_639}")
    print("Готово: found_any.csv, not_found_any.csv, only_639_excel.csv")



if __name__ == "__main__":
    main()
    
# PATH = r"C:\Users\Sergey\MAGA\Kimch\НМЦК_сборка\44-фз\Закупка № 0108500000425000045 Прил№3_Обоснование НМЦК.xlsx"
# import pandas as pd, re
# dfs = pd.read_excel(PATH, sheet_name=None, dtype=str)
# hits = []
# for name, df in dfs.items():
#     for s in df.astype(str).values.ravel():
#         s = (s or "").replace("\u00A0", " ")
#         m = RE_ART_ANY.search(s)
#         if m:
#             hits.append((name, s[max(0,m.start()-40):m.end()+40]))
#             break
# print("Нашли в листах:", [h[0] for h in hits])
# for sheet, snip in hits[:3]:
#     print(f"[{sheet}] ...{snip}...")
